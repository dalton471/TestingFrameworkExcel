"""
main.py
--------
Entry point of the KPI Validation Automation Framework.

This file does NOT contain any validation logic itself. Its only job
is to ORCHESTRATE the pipeline, in this order:

    1. Load the JSON configuration file.
    2. Load the input file (single Excel/CSV file, or every supported
       file inside an input folder).
    3. Call every validation function (from validations/validations.py).
    4. Collect all the results together.
    5. Generate the Validation_Report.xlsx (via utils/report_generator.py).

All paths are supplied on the command line via argparse:

    python main.py --config config/Kpi_automation_phase1.json --input input/KPI.xlsx
    python main.py --config config/Kpi_automation_phase1.json --input input/ --output output/Validation_Report.xlsx
    python main.py --config config/rules.json --input input/data.csv

Run "python main.py --help" to see all available options.
"""

import os
import sys
import json
import argparse
import traceback
import gc
import time
import tempfile
import shutil

import pandas as pd
import openpyxl
from openpyxl import Workbook

from utils.helper import load_json
from utils.excel_reader import load_excel
from utils.report_generator import generate_report

from validations.validations import (
    validate_sheet_list,
    validate_sheet_columns,
    validate_duplicate_check,
    validate_null_check,
    validate_formula_check,
    validate_numeric_precision,
    validate_business_rule,
    validate_reconciliation,
    validate_cross_sheet,
    validate_trend,
)


# --------------------------------------------------------------------------
# Constants
# --------------------------------------------------------------------------

SUPPORTED_EXTENSIONS = (".xlsx", ".xls", ".csv")
DEFAULT_OUTPUT_PATH = os.path.join("output", "Validation_Report.xlsx")

# Each validator is paired with the config section(s) it needs, so a
# missing/invalid section only disables THAT validator instead of
# aborting the whole file's report.
VALIDATOR_STEPS = [
    ("sheetlist", "validate_sheet_list", "Sheet Validation"),
    ("sheetvalidation", "validate_sheet_columns", "Field Validation"),
    ("dataqualityvalidation.duplicatecheck", "validate_duplicate_check", "Duplicate Check"),
    ("dataqualityvalidation.nullcheck", "validate_null_check", "Null Check"),
    ("formulavalidation", "validate_formula_check", "Formula Validation"),
    ("dataqualityvalidation.numericprecision", "validate_numeric_precision", "Numeric Precision"),
    ("businessrulevalidation", "validate_business_rule", "Business Rule Validation"),
    ("reconciliationvalidation", "validate_reconciliation", "Reconciliation Validation"),
    ("crosssheetvalidation", "validate_cross_sheet", "Cross Sheet Validation"),
    ("trendvalidation", "validate_trend", "Trend Validation"),
]


# --------------------------------------------------------------------------
# Argument Parsing
# --------------------------------------------------------------------------

def parse_arguments():
    parser = argparse.ArgumentParser(
        description="KPI Validation Automation Framework - validates "
                     "Excel/CSV data against a JSON rule set and "
                     "produces a Validation_Report.xlsx"
    )
    parser.add_argument(
        "--config", required=True,
        help="Path to the JSON configuration file "
             "(e.g. config/Kpi_automation_phase1.json)"
    )
    parser.add_argument(
        "--input", required=True,
        help="Path to a single input file (.xlsx, .xls, .csv) OR a "
             "folder containing multiple such files "
             "(e.g. input/KPI.xlsx or input/)"
    )
    parser.add_argument(
        "--output", required=False, default=DEFAULT_OUTPUT_PATH,
        help=f"Path for the generated report. Defaults to '{DEFAULT_OUTPUT_PATH}'"
    )
    return parser.parse_args()


# --------------------------------------------------------------------------
# Input Validation / Discovery Helpers
# --------------------------------------------------------------------------

def validate_config_path(config_path):
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Config file not found: '{config_path}'")
    if not os.path.isfile(config_path):
        raise ValueError(f"Config path is not a file: '{config_path}'")


def validate_input_path(input_path):
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"Input path not found: '{input_path}'")


def collect_input_files(input_path):
    """
    Accept either:
    1. A single supported input file
    2. A folder containing one or more supported input files

    Folder discovery is recursive, so files inside subfolders are also
    discovered.
    """

    if os.path.isfile(input_path):
        ext = os.path.splitext(input_path)[1].lower()

        if ext not in SUPPORTED_EXTENSIONS:
            raise ValueError(
                f"Unsupported file type '{ext}' for '{input_path}'. "
                f"Supported types are: {', '.join(SUPPORTED_EXTENSIONS)}"
            )

        return [input_path]

    if os.path.isdir(input_path):
        found_files = []

        for root, _, filenames in os.walk(input_path):
            for filename in sorted(filenames):

                # Ignore temporary Excel lock files
                if filename.startswith("~$"):
                    continue

                ext = os.path.splitext(filename)[1].lower()

                if ext in SUPPORTED_EXTENSIONS:
                    found_files.append(
                        os.path.join(root, filename)
                    )

        found_files.sort()

        if not found_files:
            raise ValueError(
                f"No supported files ({', '.join(SUPPORTED_EXTENSIONS)}) "
                f"found inside folder: '{input_path}'"
            )

        return found_files

    raise ValueError(f"Invalid input path: '{input_path}'")

def combine_split_workbook(input_folder):
    """
    Combines multiple Excel files from a folder into one temporary
    multi-sheet Excel workbook.

    Each Excel file represents one sheet:
        Read Me.xlsx              -> Read Me
        Data Pipeline Status.xlsx -> Data Pipeline Status
        KPI.xlsx                   -> KPI
        etc.

    The original validation functions continue to receive a normal
    multi-sheet Excel workbook.
    """

    if not os.path.isdir(input_folder):
        raise ValueError(
            f"Split workbook path must be a folder: '{input_folder}'"
        )

    excel_files = []

    for filename in sorted(os.listdir(input_folder)):
        if filename.startswith("~$"):
            continue

        ext = os.path.splitext(filename)[1].lower()

        if ext in (".xlsx", ".xls"):
            excel_files.append(
                os.path.join(input_folder, filename)
            )

    if not excel_files:
        raise ValueError(
            f"No Excel files found in split workbook folder: '{input_folder}'"
        )

    # Create the temporary combined workbook in a real OS temp directory
    # (never inside the output/ folder, so it can never be left behind
    # there even if deletion is delayed or retried).
    temp_dir = tempfile.mkdtemp(prefix="kpi_framework_")

    folder_name = os.path.basename(os.path.normpath(input_folder))
    temp_path = os.path.join(
        temp_dir,
        f"temp{folder_name}.xlsx"
    )

    workbook = Workbook()

    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for file_path in excel_files:

        file_name = os.path.splitext(
            os.path.basename(file_path)
        )[0]

        # Read the first/only sheet from the individual workbook using
        # openpyxl directly (not pandas). This preserves each cell's
        # number_format (e.g. "#,##0.00") alongside its value, which
        # pandas' read_excel() discards. Numeric Precision validation
        # relies on the cell's display format to know how many decimals
        # Excel actually shows - losing it here would make derived
        # (full float-precision) columns look like precision failures
        # even when the source workbook displays them correctly.
        try:
            source_workbook = openpyxl.load_workbook(file_path, data_only=True)
            source_sheet = source_workbook.active
        except Exception as error:
            raise ValueError(
                f"Could not read input file '{file_path}': {error}"
            )

        # Excel sheet names cannot exceed 31 characters
        sheet_name = file_name[:31]

        # Create sheet
        ws = workbook.create_sheet(title=sheet_name)

        # Copy every cell (header row and data rows alike), value and
        # number_format together, so the temp workbook is a faithful
        # copy of the source file rather than just its raw values.
        for row_index, source_row in enumerate(source_sheet.iter_rows(), start=1):
            for col_index, source_cell in enumerate(source_row, start=1):
                dest_cell = ws.cell(
                    row=row_index,
                    column=col_index,
                    value=source_cell.value
                )
                dest_cell.number_format = source_cell.number_format

        source_workbook.close()

    workbook.save(temp_path)

    print(
        f"Created temporary combined workbook: {temp_path}"
    )

    print("Sheets combined:")

    for file_path in excel_files:
        sheet_name = os.path.splitext(
            os.path.basename(file_path)
        )[0]

        print(f"  - {sheet_name}")

    return temp_path, temp_dir

def is_split_workbook_folder(input_path):
    """
    Returns True when the input path is a folder containing
    multiple Excel files representing individual sheets.
    """

    if not os.path.isdir(input_path):
        return False

    excel_files = []

    for filename in os.listdir(input_path):

        if filename.startswith("~$"):
            continue

        ext = os.path.splitext(filename)[1].lower()

        if ext in (".xlsx", ".xls"):
            excel_files.append(filename)

    return len(excel_files) > 1


def prepare_excel_path(file_path):
    """
    Excel files pass through unchanged. A CSV is converted into a
    temporary single-sheet .xlsx so it can flow through the same
    validation functions. A CSV has no worksheets, so every
    sheet-based/cross-sheet rule will correctly report a missing
    sheet for anything other than that one logical sheet - this is
    expected, not a bug (documented in the README).
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext in (".xlsx", ".xls"):
        return file_path, False

    if ext == ".csv":
        try:
            df = pd.read_csv(file_path)
        except Exception as error:
            raise ValueError(f"Could not read CSV file '{file_path}': {error}")

        os.makedirs("output", exist_ok=True)
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        temp_path = os.path.join("output", f"_temp_{base_name}.xlsx")

        try:
            df.to_excel(temp_path, index=False, sheet_name="Sheet1")
        except Exception as error:
            raise ValueError(f"Could not convert CSV to Excel for validation: {error}")

        return temp_path, True

    raise ValueError(f"Unsupported file type '{ext}' for '{file_path}'")


def build_output_path_for_file(output_arg, input_file_path, is_batch):
    if not is_batch:
        return output_arg

    output_dir = os.path.dirname(output_arg) or "."
    base, ext = os.path.splitext(os.path.basename(output_arg))
    input_stem = os.path.splitext(os.path.basename(input_file_path))[0]
    return os.path.join(output_dir, f"{base}_{input_stem}{ext}")


def _get_config_section(config, dotted_path):
    """Walk a dotted path like 'dataqualityvalidation.nullcheck' through
    the config dict. Raises KeyError with the full path if missing."""
    node = config
    for part in dotted_path.split("."):
        node = node[part]  # KeyError propagates with the missing part
    return node


# --------------------------------------------------------------------------
# Core Orchestration
# --------------------------------------------------------------------------

VALIDATOR_FUNCTIONS = {
    "validate_sheet_list": validate_sheet_list,
    "validate_sheet_columns": validate_sheet_columns,
    "validate_duplicate_check": validate_duplicate_check,
    "validate_null_check": validate_null_check,
    "validate_formula_check": validate_formula_check,
    "validate_numeric_precision": validate_numeric_precision,
    "validate_business_rule": validate_business_rule,
    "validate_reconciliation": validate_reconciliation,
    "validate_cross_sheet": validate_cross_sheet,
    "validate_trend": validate_trend,
}

def print_validation_results(results):
    """Print a complete validation summary to the terminal."""

    print("\n" + "=" * 80)
    print("VALIDATION RESULTS")
    print("=" * 80)

    for result in results:
        sheet_name = result.get("Sheet Name", "")
        field = result.get("Field", "")
        test_name = result.get("Test Name", "")
        status = result.get("Status (P/F)", "")
        failed_count = result.get("Failed Count", 0)

        if status == "P":
            print(
                f"{sheet_name} | {field} | "
                f"{test_name} | PASS | Failed Count: {failed_count}"
            )
        else:
            print(
                f"{sheet_name} | {field} | "
                f"{test_name} | FAIL | Failed Count: {failed_count}"
            )

    print("=" * 80)
    print("VALIDATION COMPLETED")
    print("=" * 80)

def run_validation_pipeline(config, excel_file, output_file):
    """
    Runs every validator. A missing/invalid config section for ONE
    validator does not abort the others - it is recorded as a single
    "Configuration Error" row and the rest of the pipeline continues.
    """

    workbook = None

    try:
        workbook = load_excel(excel_file)
        print("Excel Loaded Successfully")

        final_results = []

        for config_path, func_name, test_name in VALIDATOR_STEPS:
            func = VALIDATOR_FUNCTIONS[func_name]

            try:
                _get_config_section(config, config_path)

                if func_name == "validate_sheet_list":
                    section_results = func(config, workbook)
                else:
                    section_results = func(config, excel_file)

                final_results.extend(section_results)

            except KeyError as missing_key:
                print(
                    f"[CONFIG ERROR] Missing configuration section "
                    f"'{config_path}' ({missing_key}) - skipping {test_name}."
                )

                final_results.append({
                    "Sheet Name": "Configuration",
                    "Field": config_path,
                    "Test Type": "Configuration Validation",
                    "Test Name": test_name,
                    "Status (P/F)": "F",
                    "Failed Count": 1,
                })

            except Exception as error:
                print(
                    f"[VALIDATOR ERROR] {test_name} failed: {error}"
                )
                traceback.print_exc()

                final_results.append({
                    "Sheet Name": "Configuration",
                    "Field": test_name,
                    "Test Type": "Configuration Validation",
                    "Test Name": test_name,
                    "Status (P/F)": "F",
                    "Failed Count": 1,
                })

        # Print all validation results to terminal
        print_validation_results(final_results)

        output_dir = os.path.dirname(output_file)

        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        generate_report(final_results, output_file)

    finally:
        # Close the pandas ExcelFile so Windows can release the file lock.
        if workbook is not None:
            try:
                workbook.close()
                print("Input workbook closed successfully")
            except Exception as error:
                print(
                    f"[WARNING] Could not close input workbook: {error}"
                )
                
def process_single_file(config, input_file, output_file):
    temp_path = None

    try:
        print(f"\n{'=' * 60}")
        print(f"Processing: {input_file}")
        print(f"{'=' * 60}")

        excel_path, is_temporary = prepare_excel_path(input_file)
        if is_temporary:
            temp_path = excel_path

        run_validation_pipeline(config, excel_path, output_file)

        print(f"Report generated: {output_file}")
        return True

    except FileNotFoundError as error:
        print(f"[FAILED] File not found while processing '{input_file}': {error}")
        return False

    except ValueError as error:
        print(f"[FAILED] Invalid file/path while processing '{input_file}': {error}")
        return False

    except (pd.errors.ParserError, pd.errors.EmptyDataError) as error:
        print(f"[FAILED] Invalid/corrupt Excel or CSV content in '{input_file}': {error}")
        return False

    except Exception as error:
        print(f"[FAILED] Unexpected error while processing '{input_file}': {error}")
        traceback.print_exc()
        return False

    finally:
        if temp_path and os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass


# --------------------------------------------------------------------------
# Main Entry Point
# --------------------------------------------------------------------------

def main():
    args = parse_arguments()

    try:
        validate_config_path(args.config)
        validate_input_path(args.input)
    except (FileNotFoundError, ValueError) as error:
        print(f"[ERROR] {error}")
        sys.exit(1)

    try:
        config = load_json(args.config)
        print("JSON Loaded Successfully")
    except json.JSONDecodeError as error:
        print(f"[ERROR] Config file is not valid JSON: '{args.config}' -> {error}")
        sys.exit(1)
    except FileNotFoundError:
        print(f"[ERROR] Config file not found: '{args.config}'")
        sys.exit(1)
    except Exception as error:
        print(f"[ERROR] Unexpected error while loading config: {error}")
        traceback.print_exc()
        sys.exit(1)

    if not isinstance(config, dict) or not config:
        print(f"[ERROR] Config file '{args.config}' is empty or not a JSON object.")
        sys.exit(1)

    # --------------------------------------------------------------
    # Determine input mode
    # --------------------------------------------------------------

    is_split_workbook = is_split_workbook_folder(args.input)
    combined_input = None
    combined_temp_dir = None

    if is_split_workbook:

        print("\n" + "=" * 60)
        print("SPLIT WORKBOOK INPUT DETECTED")
        print("=" * 60)
        print(f"Input folder: {args.input}")

        try:
            combined_input, combined_temp_dir = combine_split_workbook(args.input)
        except Exception as error:
            print(f"[ERROR] Could not combine input files: {error}")
            traceback.print_exc()
            sys.exit(1)

        input_files = [combined_input]
        is_batch = False

    else:

        try:
            input_files = collect_input_files(args.input)
        except ValueError as error:
            print(f"[ERROR] {error}")
            sys.exit(1)

        is_batch = len(input_files) > 1

        if is_batch:
            print(
                f"Found {len(input_files)} supported file(s) "
                f"in folder '{args.input}'"
            )

    # --------------------------------------------------------------
    # Process input files
    # --------------------------------------------------------------

    success_count = 0
    failure_count = 0

    try:
        for input_file in input_files:

            output_file = build_output_path_for_file(
                args.output,
                input_file,
                is_batch
            )

            succeeded = process_single_file(
                config,
                input_file,
                output_file
            )

            if succeeded:
                success_count += 1
            else:
                failure_count += 1

    finally:
        # --------------------------------------------------------------
        # Remove the temporary combined workbook - this runs even if an
        # unexpected error occurs above, so the temp directory is never
        # left behind. It was created in a real OS temp directory (not
        # inside output/), so even if removal is briefly delayed, it
        # can never show up alongside Validation_Report.xlsx.
        # --------------------------------------------------------------

        if is_split_workbook and combined_temp_dir:

            # Force garbage collection before removing the file so any
            # remaining workbook/file handles (pandas ExcelFile,
            # openpyxl workbooks opened for cell-format inspection,
            # etc.) are released on Windows.
            gc.collect()

            if os.path.exists(combined_temp_dir):
                removed = False

                # Windows can briefly keep the workbook locked after the
                # ExcelFile is closed. Retry a few times before warning.
                for attempt in range(5):
                    try:
                        shutil.rmtree(combined_temp_dir)
                        removed = True

                        print(
                            f"Temporary combined workbook removed: "
                            f"{combined_input}"
                        )
                        break

                    except PermissionError:
                        if attempt < 4:
                            gc.collect()
                            time.sleep(1)

                    except OSError as error:
                        print(
                            f"[WARNING] Could not remove temporary workbook "
                            f"directory: {error}"
                        )
                        break

                if not removed and os.path.exists(combined_temp_dir):
                    print(
                        f"[WARNING] Temporary combined workbook is still in "
                        f"use and could not be removed: {combined_temp_dir}"
                    )

    # --------------------------------------------------------------
    # Run Summary
    # --------------------------------------------------------------

    print(f"\n{'=' * 60}")
    print("Run Summary")
    print(f"{'=' * 60}")
    print(f"Total files:  {len(input_files)}")
    print(f"Succeeded:    {success_count}")
    print(f"Failed:       {failure_count}")

    if success_count == 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
