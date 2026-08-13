"""
validations.py
----------------
Generic, configuration-driven validation engines for the KPI framework.

Every function reads its parameters from the JSON configuration (sheet
names, columns, operators, thresholds, precision, lookup relationships,
tolerances, blank-value indicators, etc). None of these functions
contain business-rule names, practice codes, or expected counts -
those all come from the config and the data.

Functions:
    - validate_sheet_list()
    - validate_sheet_columns()
    - validate_duplicate_check()
    - validate_null_check()
    - validate_formula_check()
    - validate_numeric_precision()
    - validate_business_rule()      (generic engine, dispatches on rule["type"])
    - validate_reconciliation()
    - validate_cross_sheet()
    - validate_trend()
"""

import math

import pandas as pd

try:
    import openpyxl
except ImportError:  # pragma: no cover - openpyxl is a hard dependency
    openpyxl = None


# --------------------------------------------------------------------------
# Small shared helpers
# --------------------------------------------------------------------------

def _clean_column_name(name):
    """Lower-case, strip spaces/underscores - used only for loose matching
    of column *existence*, never for reading actual values."""
    return str(name).strip().lower().replace(" ", "").replace("_", "")


def _normalize_text(value, blank_indicators=None):
    """Return a trimmed, lower-cased string form of a value. NaN/None
    becomes an empty string. `blank_indicators` (from config) is a list
    of strings that should also be treated as "blank"."""
    if value is None or (isinstance(value, float) and pd.isna(value)) or pd.isna(value):
        text = ""
    else:
        text = str(value).strip().lower()

    blank_indicators = [b.strip().lower() for b in (blank_indicators or [])]
    if text in blank_indicators:
        text = ""
    return text


def _read_sheet(excel_file, sheet_name):
    """Read a sheet, returning (df, error). df is None if the sheet
    could not be read (missing sheet, corrupt file, etc)."""
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        return df, None
    except Exception as error:
        return None, str(error)


def _missing_sheet_result(sheet_name, test_type, test_name, field="Sheet"):
    return {
        "Sheet Name": sheet_name,
        "Field": field,
        "Test Type": test_type,
        "Test Name": test_name,
        "Status (P/F)": "F",
        "Failed Count": 1,
    }


def _missing_columns_result(sheet_name, columns, test_type, test_name):
    return {
        "Sheet Name": sheet_name,
        "Field": ", ".join(columns),
        "Test Type": test_type,
        "Test Name": test_name,
        "Status (P/F)": "F",
        "Failed Count": len(columns),
    }


def _apply_operator(left_value, operator, right_value):
    """Evaluate left_value OPERATOR right_value using a configured
    operator string. Supports the standard comparison operators."""
    operators = {
        "<": lambda a, b: a < b,
        "<=": lambda a, b: a <= b,
        ">": lambda a, b: a > b,
        ">=": lambda a, b: a >= b,
        "==": lambda a, b: a == b,
        "=": lambda a, b: a == b,
        "!=": lambda a, b: a != b,
    }
    if operator not in operators:
        raise ValueError(f"Unsupported operator: {operator}")
    return operators[operator](left_value, right_value)


# --------------------------------------------------------------------------
# 1. Sheet existence
# --------------------------------------------------------------------------

def validate_sheet_list(config, workbook):
    results = []
    json_sheets = config["sheetlist"]
    excel_sheets = workbook.sheet_names

    for sheet in json_sheets:
        status = "P" if sheet in excel_sheets else "F"
        results.append({
            "Sheet Name": sheet,
            "Field": "Sheet",
            "Test Type": "Sheet Validation",
            "Test Name": "Sheet Existence Testing",
            "Status (P/F)": status,
            "Failed Count": 0 if status == "P" else 1,
        })
    return results


# --------------------------------------------------------------------------
# 2. Column / field existence
# --------------------------------------------------------------------------

def validate_sheet_columns(config, excel_file):
    results = []
    sheet_validation = config["sheetvalidation"]

    for sheet_name, sheet_info in sheet_validation.items():
        df, error = _read_sheet(excel_file, sheet_name)

        if df is None:
            results.append(_missing_sheet_result(
                sheet_name, "Field Validation", "Field Existence Testing", field="Field"))
            continue

        excel_columns = {_clean_column_name(c) for c in df.columns}

        for field in sheet_info["fields"]:
            print("DEBUG FIELD:", sheet_name, field["name"])

            json_column = _clean_column_name(field["name"])
            status = "P" if json_column in excel_columns else "F"

            results.append({
                "Sheet Name": sheet_name,
                "Field": field["name"],
                "Test Type": "Field Validation",
                "Test Name": "Field Existence Testing",
                "Status (P/F)": status,
                "Failed Count": 0 if status == "P" else 1,
            })
    return results


# --------------------------------------------------------------------------
# 3. Duplicate check
# --------------------------------------------------------------------------

def validate_duplicate_check(config, excel_file):
    results = []
    rules = config["dataqualityvalidation"]["duplicatecheck"]

    for rule in rules:
        sheet_name = rule["sheetname"]
        columns = rule["columns"]

        df, error = _read_sheet(excel_file, sheet_name)
        if df is None:
            results.append(_missing_sheet_result(
                sheet_name, "Data Quality Validation", "Duplicate Check"))
            continue

        missing_columns = [c for c in columns if c not in df.columns]
        if missing_columns:
            results.append(_missing_columns_result(
                sheet_name, missing_columns, "Data Quality Validation", "Duplicate Check"))
            continue

        # Null/empty keys are not meaningful duplicate groups on their own -
        # exclude rows where every key column is null before checking.
        key_df = df[columns]
        non_null_mask = ~key_df.isnull().all(axis=1)
        candidate_rows = df[non_null_mask]

        duplicate_rows = candidate_rows[candidate_rows.duplicated(subset=columns, keep=False)]
        duplicate_count = len(duplicate_rows)
        status = "P" if duplicate_count == 0 else "F"

        results.append({
            "Sheet Name": sheet_name,
            "Field": ", ".join(columns),
            "Test Type": "Data Quality Validation",
            "Test Name": "Duplicate Check",
            "Status (P/F)": status,
            "Failed Count": duplicate_count,
        })
    return results


# --------------------------------------------------------------------------
# 4. Null check
# --------------------------------------------------------------------------

def validate_null_check(config, excel_file):
    results = []
    rules = config["dataqualityvalidation"]["nullcheck"]

    for rule in rules:
        sheet_name = rule["sheetname"]
        columns = rule["columns"]

        df, error = _read_sheet(excel_file, sheet_name)
        if df is None:
            results.append(_missing_sheet_result(
                sheet_name, "Data Quality Validation", "Null Check"))
            continue

        status = "P"
        failed_count = 0

        for column in columns:
            if column not in df.columns:
                status = "F"
                failed_count += 1
                continue

            # Treat None/NaN/empty-string/whitespace-only as null.
            series = df[column]
            is_blank = series.isna() | series.astype(str).str.strip().eq("")
            null_count = int(is_blank.sum())

            if null_count > 0:
                status = "F"
                failed_count += null_count

        results.append({
            "Sheet Name": sheet_name,
            "Field": ", ".join(columns),
            "Test Type": "Data Quality Validation",
            "Test Name": "Null Check",
            "Status (P/F)": status,
            "Failed Count": failed_count,
        })
    return results


# --------------------------------------------------------------------------
# 5. Formula validation (threshold / comparison, JSON-defined)
# --------------------------------------------------------------------------

def validate_formula_check(config, excel_file):
    results = []
    formula_rules = config["formulavalidation"]

    df, error = _read_sheet(excel_file, "KPI")
    if df is None:
        return [_missing_sheet_result("KPI", "Data Quality Validation", "Formula Validation")]

    df.columns = df.columns.str.strip().str.replace(" ", "_").str.replace("/", "_")

    for rule in formula_rules:
        formula_name = rule["name"]
        formula = rule["formula"]
        validation_type = rule["validationtype"].lower()
        aggregate = rule.get("aggregate", False)

        try:
            if validation_type == "threshold":
                threshold = rule["threshold"]
                operator = rule["operator"]

                if aggregate:
                    group_columns = [c.strip().replace(" ", "_") for c in rule["groupbycolumns"]]
                    agg_dict = {c.strip().replace(" ", "_"): "sum" for c in rule.get("sourcecolumns", [])}
                    grouped_df = df.groupby(group_columns, as_index=False).agg(agg_dict)
                    calculated = grouped_df.eval(formula)
                else:
                    calculated = df.eval(formula)

                failed_mask = (calculated.notna() & ~calculated.apply(lambda v: _apply_operator(v, operator, threshold) if pd.notna(v) else False))
                failed_count = int(failed_mask.sum())

            elif validation_type == "comparison":
                target_column = rule["targetcolumn"].strip().replace(" ", "_")
                source_columns = [c.strip().replace(" ", "_") for c in rule["sourcecolumns"]]
                precision = rule.get("precision", 2)

                if aggregate:
                    group_columns = [c.strip().replace(" ", "_") for c in rule["groupbycolumns"]]
                    agg_dict = {c: "sum" for c in source_columns}
                    agg_dict[target_column] = "sum"
                    grouped_df = df.groupby(group_columns, as_index=False).agg(agg_dict)
                else:
                    grouped_df = df.copy()

                calculated = grouped_df.eval(formula)
                target_values = grouped_df[target_column]

                failed_mask = (calculated.round(precision) != target_values.round(precision))
                failed_count = int(failed_mask.sum())

            else:
                raise ValueError(f"Unsupported validation type: {validation_type}")

            status = "P" if failed_count == 0 else "F"

        except Exception as e:
            status = "F"
            failed_count = 1
            print(f"{formula_name} - FAIL ({str(e)})")

        results.append({
            "Sheet Name": "KPI",
            "Field": formula_name,
            "Test Type": "Data Quality Validation",
            "Test Name": "Formula Validation",
            "Status (P/F)": status,
            "Failed Count": failed_count,
        })
    return results


# --------------------------------------------------------------------------
# 6. Numeric precision (format-aware root-cause fix)
# --------------------------------------------------------------------------

def _format_decimal_places(number_format):
    """Given an Excel cell number_format string, return how many decimal
    digits it *displays*, or None if the format doesn't force a fixed
    number of decimals (e.g. 'General' or a text/date format)."""
    if not number_format or number_format == "General":
        return None

    # Only look at the first (positive-number) section of the format.
    section = number_format.split(";")[0]

    if "." not in section:
        return 0

    frac_part = section.split(".")[-1]
    count = 0
    for ch in frac_part:
        if ch in "0#":
            count += 1
        else:
            break
    return count


def _raw_value_decimal_places(value, precision):
    """Fallback used when a cell has no explicit numeric format (e.g.
    'General', or the workbook can't be opened with openpyxl - a CSV).
    Uses a tolerance so float rounding noise (…0000000004) isn't
    mistaken for genuine extra precision."""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return 0

    if not math.isfinite(v):
        return 0

    rounded = round(v, precision)
    if abs(v - rounded) <= 1e-6:
        return precision  # effectively equal once rounded -> passes

    # Count genuine significant decimal digits (bounded to avoid float
    # repr artifacts beyond double precision).
    text = f"{v:.10f}".rstrip("0")
    if "." in text:
        return len(text.split(".")[1])
    return 0


def _build_format_lookup(excel_file, sheet_name, columns):
    """Return {column_name: [number_format, ...]} for each row's cell in
    that column, using openpyxl. Returns {} if it can't be opened
    (e.g. CSV-derived xlsx without preserved styles, or old .xls)."""
    if openpyxl is None:
        return {}
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        if sheet_name not in wb.sheetnames:
            return {}
        ws = wb[sheet_name]
    except Exception:
        return {}

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header_index = {}
    for idx, cell in enumerate(header_row):
        if cell.value is not None:
            header_index[str(cell.value).strip()] = idx

    wanted = {col: header_index[col] for col in columns if col in header_index}
    if not wanted:
        return {}

    formats = {col: [] for col in wanted}
    for row in ws.iter_rows(min_row=2):
        for col, idx in wanted.items():
            if idx < len(row):
                cell = row[idx]
                formats[col].append((cell.value, cell.number_format))
    return formats


def validate_numeric_precision(config, excel_file):
    results = []
    rules = config["dataqualityvalidation"]["numericprecision"]

    for rule in rules:
        sheet_name = rule["sheetname"]
        precision = rule["precision"]
        use_cell_format = rule.get("usecellformat", True)

        df, error = _read_sheet(excel_file, sheet_name)
        if df is None:
            results.append(_missing_sheet_result(
                sheet_name, "Data Quality Validation", "Numeric Precision"))
            continue

        if rule["columns"] == "ALL_NUMERIC_COLUMNS":
            numeric_columns = list(df.select_dtypes(include="number").columns)
        else:
            numeric_columns = list(rule["columns"])

        if not numeric_columns:
            results.append({
                "Sheet Name": sheet_name,
                "Field": "Numeric Columns",
                "Test Type": "Data Quality Validation",
                "Test Name": "Numeric Precision",
                "Status (P/F)": "F",
                "Failed Count": 1,
            })
            continue

        format_lookup = _build_format_lookup(excel_file, sheet_name, numeric_columns) if use_cell_format else {}

        for column in numeric_columns:
            if column not in df.columns:
                results.append(_missing_columns_result(
                    sheet_name, [column], "Data Quality Validation", "Numeric Precision"))
                continue

            failed_count = 0

            if column in format_lookup:
                for value, number_format in format_lookup[column]:
                    if value is None or (isinstance(value, float) and pd.isna(value)):
                        continue
                    if not isinstance(value, (int, float)):
                        continue

                    fmt_places = _format_decimal_places(number_format)
                    if fmt_places is not None:
                        # The workbook itself defines how many decimals are
                        # ever displayed for this cell - trust that.
                        if fmt_places > precision:
                            failed_count += 1
                    else:
                        # No explicit format ('General') - fall back to the
                        # tolerant raw-value check.
                        if _raw_value_decimal_places(value, precision) > precision:
                            failed_count += 1
            else:
                # openpyxl/styles unavailable (e.g. CSV input) - fall back
                # to the tolerant raw-value check for every value.
                for value in df[column].dropna():
                    if _raw_value_decimal_places(value, precision) > precision:
                        failed_count += 1

            status = "P" if failed_count == 0 else "F"
            results.append({
                "Sheet Name": sheet_name,
                "Field": column,
                "Test Type": "Data Quality Validation",
                "Test Name": "Numeric Precision",
                "Status (P/F)": status,
                "Failed Count": failed_count,
            })
    return results


# --------------------------------------------------------------------------
# 7. Business rule validation - generic engine, dispatched by rule["type"]
# --------------------------------------------------------------------------

def _business_rule_comparison(rule, df):
    left_column = rule["leftcolumn"]
    right_column = rule["rightcolumn"]
    distinct_column = rule.get("distinctcolumn")
    operator = rule["operator"]
    datatype = rule.get("datatype", "date")
    blank_indicators = rule.get("blankindicators", [])

    failed_rows = []

    for _, row in df.iterrows():
        left_raw, right_raw = row[left_column], row[right_column]
        left_text = _normalize_text(left_raw, blank_indicators)
        right_text = _normalize_text(right_raw, blank_indicators)

        if left_text == "" and right_text == "":
            failed_rows.append(row)  # both blank: not a valid comparison state
            continue
        if (left_text == "") != (right_text == ""):
            failed_rows.append(row)  # only one side blank: inconsistent
            continue

        if datatype == "date":
            left_val = pd.to_datetime(left_raw, errors="coerce")
            right_val = pd.to_datetime(right_raw, errors="coerce")
        else:
            left_val = pd.to_numeric(left_raw, errors="coerce")
            right_val = pd.to_numeric(right_raw, errors="coerce")

        if pd.isna(left_val) or pd.isna(right_val):
            failed_rows.append(row)
            continue

        if not _apply_operator(left_val, operator, right_val):
            failed_rows.append(row)

    failed_df = pd.DataFrame(failed_rows)
    if distinct_column and not failed_df.empty and distinct_column in failed_df.columns:
        failed_df = failed_df.drop_duplicates(subset=[distinct_column])

    return len(failed_df)


def _business_rule_staledata(rule, df):
    column_name = rule["column"]
    reference_column = rule["referencecolumn"]
    months = rule["months"]
    distinct_column = rule.get("distinctcolumn")
    blank_indicators = rule.get("blankindicators", [])

    original_text = df[column_name].apply(lambda v: _normalize_text(v, blank_indicators))

    parsed_column = pd.to_datetime(df[column_name], errors="coerce")

    # A transaction date older than the configured number of months
    # from today must be displayed as NULL / Not Applicable.
    threshold_date = pd.Timestamp.today().normalize() - pd.DateOffset(months=months)

    is_stale_but_shown = (
        parsed_column.notna()
        & (parsed_column < threshold_date)
        & (original_text != "")
    )

    failed_df = df.loc[is_stale_but_shown]
    if distinct_column and not failed_df.empty and distinct_column in failed_df.columns:
        failed_df = failed_df.drop_duplicates(subset=[distinct_column])

    return len(failed_df)


def _business_rule_lookupmatch(rule, df, excel_file):
    left_column = rule["leftcolumn"]
    lookup_sheet = rule["lookupsheet"]
    lookup_key_left = rule["lookupkeyleft"]
    lookup_key_right = rule["lookupkeyright"]
    right_column = rule["rightcolumn"]
    drop_null_keys = rule.get("dropnullkeys", True)

    lookup_df, error = _read_sheet(excel_file, lookup_sheet)
    if lookup_df is None:
        return None, f"Lookup sheet not found: {lookup_sheet}"

    missing_lookup_columns = [c for c in (lookup_key_right, right_column) if c not in lookup_df.columns]
    if missing_lookup_columns:
        return None, f"Missing lookup column(s): {', '.join(missing_lookup_columns)}"

    working_df = df
    if drop_null_keys:
        working_df = working_df[working_df[lookup_key_left].notna()]

    merged_df = working_df.merge(
        lookup_df[[lookup_key_right, right_column]],
        left_on=lookup_key_left, right_on=lookup_key_right, how="left",
        suffixes=("", "_lookup"),
    )

    left_text = merged_df[left_column].apply(_normalize_text)
    right_text = merged_df[right_column].apply(_normalize_text)

    mismatch = left_text != right_text
    failed_df = merged_df.loc[mismatch]
    failed_df = failed_df.drop_duplicates(subset=[lookup_key_left])

    return len(failed_df), None


def validate_business_rule(config, excel_file):
    results = []
    business_rules = config["businessrulevalidation"]

    for rule in business_rules:
        rule_name = rule["name"]
        sheet_name = rule["sheetname"]
        rule_type = rule.get("type")

        df, error = _read_sheet(excel_file, sheet_name)
        if df is None:
            results.append(_missing_sheet_result(
                sheet_name, "Data Validation", "Business Rule Validation", field=rule_name))
            continue

        # Only the columns that must live on the PRIMARY sheet. For
        # lookupmatch rules, rightcolumn/lookupkeyright belong to the
        # lookup sheet and are validated separately.
        primary_columns = []
        for key in ("leftcolumn", "column", "referencecolumn", "distinctcolumn", "lookupkeyleft"):
            if key in rule:
                primary_columns.append(rule[key])
        if rule_type != "lookupmatch" and "rightcolumn" in rule:
            primary_columns.append(rule["rightcolumn"])

        missing_columns = [c for c in primary_columns if c not in df.columns]
        if missing_columns:
            results.append(_missing_columns_result(
                sheet_name, missing_columns, "Data Validation", "Business Rule Validation"))
            continue

        try:
            if rule_type == "comparison":
                failed_count = _business_rule_comparison(rule, df)
            elif rule_type == "staledata":
                failed_count = _business_rule_staledata(rule, df)
            elif rule_type == "lookupmatch":
                failed_count, lookup_error = _business_rule_lookupmatch(rule, df, excel_file)
                if failed_count is None:
                    results.append({
                        "Sheet Name": sheet_name,
                        "Field": rule_name,
                        "Test Type": "Data Validation",
                        "Test Name": "Business Rule Validation",
                        "Status (P/F)": "F",
                        "Failed Count": 1,
                    })
                    continue
            else:
                raise ValueError(f"Unsupported business rule type: {rule_type}")

            status = "P" if failed_count == 0 else "F"

        except Exception as e:
            status = "F"
            failed_count = 1
            print(f"{rule_name} - FAIL ({str(e)})")

        results.append({
            "Sheet Name": sheet_name,
            "Field": rule_name,
            "Test Type": "Data Validation",
            "Test Name": "Business Rule Validation",
            "Status (P/F)": status,
            "Failed Count": failed_count,
        })
    return results


# --------------------------------------------------------------------------
# 8. Reconciliation validation (numeric tolerance instead of raw equality)
# --------------------------------------------------------------------------

def validate_reconciliation(config, excel_file):
    results = []
    rules = config["reconciliationvalidation"]

    for rule in rules:
        rule_name = rule["name"]
        source_sheet = rule["sourcesheet"]
        target_sheet = rule["targetsheet"]
        match_columns = [c.strip().replace(" ", "_") for c in rule["matchcolumns"]]
        source_column = rule["sourcecolumn"].strip().replace(" ", "_")
        target_column = rule["targetcolumn"].strip().replace(" ", "_")
        distinct_column = rule["distinctcolumn"].strip().replace(" ", "_")
        tolerance = rule.get("tolerance", 0.01)

        source_df, error = _read_sheet(excel_file, source_sheet)
        if source_df is None:
            results.append(_missing_sheet_result(
                source_sheet, "Data Validation", "Reconciliation Validation", field="Source Sheet"))
            continue

        target_df, error = _read_sheet(excel_file, target_sheet)
        if target_df is None:
            results.append(_missing_sheet_result(
                target_sheet, "Data Validation", "Reconciliation Validation", field="Target Sheet"))
            continue

        source_df.columns = [str(c).strip().replace(" ", "_") for c in source_df.columns]
        target_df.columns = [str(c).strip().replace(" ", "_") for c in target_df.columns]

        missing_match_columns = [c for c in match_columns if c not in source_df.columns or c not in target_df.columns]
        if missing_match_columns:
            results.append(_missing_columns_result(
                source_sheet, missing_match_columns, "Data Validation", "Reconciliation Validation"))
            continue

        if source_column not in source_df.columns:
            results.append(_missing_columns_result(
                source_sheet, [source_column], "Data Validation", "Reconciliation Validation"))
            continue

        if target_column not in target_df.columns:
            results.append(_missing_columns_result(
                target_sheet, [target_column], "Data Validation", "Reconciliation Validation"))
            continue

        source_df[source_column] = pd.to_numeric(source_df[source_column], errors="coerce").fillna(0)
        target_df[target_column] = pd.to_numeric(target_df[target_column], errors="coerce").fillna(0)

        source_agg = source_df.groupby(match_columns, as_index=False).agg({source_column: "sum"})
        target_agg = target_df.groupby(match_columns, as_index=False).agg({target_column: "sum"})

        merged_df = pd.merge(source_agg, target_agg, on=match_columns, how="outer").fillna(0)

        diff = (merged_df[source_column] - merged_df[target_column]).abs()
        failed_rows = merged_df.loc[diff > tolerance]

        failed_count = len(failed_rows)
        status = "P" if failed_count == 0 else "F"

        results.append({
            "Sheet Name": source_sheet,
            "Field": rule_name,
            "Test Type": "Data Validation",
            "Test Name": "Reconciliation Validation",
            "Status (P/F)": status,
            "Failed Count": failed_count,
        })
    return results


# --------------------------------------------------------------------------
# 9. Cross-sheet validation (dedup-by-key fix to prevent merge inflation)
# --------------------------------------------------------------------------

def validate_cross_sheet(config, excel_file):
    results = []
    rules = config.get("crosssheetvalidation", [])

    for rule in rules:
        rule_name = rule.get("name")
        source_sheet = rule.get("sourcesheet")
        target_sheet = rule.get("targetsheet")
        match_column = rule.get("matchcolumn")
        validation_type = rule.get("validationtype")
        drop_null_keys = rule.get("dropnullkeys", True)

        source_df, error = _read_sheet(excel_file, source_sheet)
        if source_df is None:
            results.append(_missing_sheet_result(
                source_sheet, "Data Validation", "Cross Sheet Validation", field="Source Sheet"))
            continue

        target_df, error = _read_sheet(excel_file, target_sheet)
        if target_df is None:
            results.append(_missing_sheet_result(
                target_sheet, "Data Validation", "Cross Sheet Validation", field="Target Sheet"))
            continue

        source_df.columns = source_df.columns.str.strip().str.lower()
        target_df.columns = target_df.columns.str.strip().str.lower()
        match_col = match_column.strip().lower()

        if match_col not in source_df.columns or match_col not in target_df.columns:
            missing_side = source_sheet if match_col not in source_df.columns else target_sheet
            results.append(_missing_columns_result(
                missing_side, [match_col], "Data Validation", "Cross Sheet Validation"))
            continue

        if drop_null_keys:
            source_df = source_df[source_df[match_col].notna()]
            target_df = target_df[target_df[match_col].notna()]

        if validation_type == "valuecomparison":
            source_column = rule.get("sourcecolumn").strip().lower()
            target_column = rule.get("targetcolumn").strip().lower()

            missing = [c for c, d, s in ((source_column, source_df, source_sheet), (target_column, target_df, target_sheet)) if c not in d.columns]
            if missing:
                results.append(_missing_columns_result(
                    source_sheet, missing, "Data Validation", "Cross Sheet Validation"))
                continue

            merged_df = pd.merge(
                source_df[[match_col, source_column]].drop_duplicates(subset=[match_col]),
                target_df[[match_col, target_column]],
                on=match_col, how="outer",
            )

            left_text = merged_df[source_column].apply(_normalize_text)
            right_text = merged_df[target_column].apply(_normalize_text)
            comparison = left_text != right_text

            failed_rows = merged_df.loc[comparison].drop_duplicates(subset=[match_col])
            failed_count = len(failed_rows)

        elif validation_type == "existencecheck":
            source_keys = set(source_df[match_col].dropna().astype(str).str.strip())
            target_keys = set(target_df[match_col].dropna().astype(str).str.strip())

            missing_keys = source_keys.symmetric_difference(target_keys)
            failed_count = len(missing_keys)

        else:
            continue

        status = "P" if failed_count == 0 else "F"

        results.append({
            "Sheet Name": source_sheet,
            "Field": rule_name,
            "Test Type": "Data Validation",
            "Test Name": "Cross Sheet Validation",
            "Status (P/F)": status,
            "Failed Count": failed_count,
        })
    return results


# --------------------------------------------------------------------------
# 10. Trend validation (respects the configured operator everywhere)
# --------------------------------------------------------------------------

def validate_trend(config, excel_file):
    results = []
    rules = config.get("trendvalidation", [])

    for rule in rules:
        sheet_name = rule["sheet"]
        metric_name = rule["name"]
        metric = rule["metric"]
        previous_metric = rule["previousmetric"]
        operator = rule["operator"]
        threshold = rule["threshold"]
        aggregate = rule.get("aggregate", False)
        groupby_columns = rule.get("groupbycolumns", [])
        distinct_column = rule.get("distinctcolumn", "PracticeCode")

        df, error = _read_sheet(excel_file, sheet_name)
        if df is None:
            results.append(_missing_sheet_result(
                sheet_name, "Data Validation", "Trend Validation", field=metric_name))
            continue

        df.columns = df.columns.str.strip().str.replace(" ", "_").str.replace("/", "_")

        if aggregate:
            required_columns = [metric] + groupby_columns
        else:
            required_columns = [metric, previous_metric, distinct_column]

        missing_columns = [c for c in required_columns if c not in df.columns]
        if missing_columns:
            results.append(_missing_columns_result(
                sheet_name, missing_columns, "Data Validation", "Trend Validation"))
            continue

        failed_count = 0

        if aggregate:
            df[metric] = pd.to_numeric(df[metric], errors="coerce").fillna(0)
            grouped = df.groupby(groupby_columns, as_index=False).agg({metric: "sum"})
            grouped = grouped.sort_values(by=groupby_columns).reset_index(drop=True)
            previous_values = grouped.groupby(groupby_columns[0])[metric].shift(1)

            for current, previous in zip(grouped[metric], previous_values):
                if pd.isna(previous) or previous == 0:
                    continue
                deviation = ((current - previous) / previous) * 100
                if _apply_operator(abs(deviation), operator, threshold):
                    failed_count += 1
        else:
            group_cols = [distinct_column] + [c for c in ("FiscalYear", "FiscalMonth") if c in df.columns]
            agg_df = df.groupby(group_cols, as_index=False).agg({metric: "sum", previous_metric: "sum"})

            current_values = pd.to_numeric(agg_df[metric], errors="coerce")
            previous_values = pd.to_numeric(agg_df[previous_metric], errors="coerce")

            for current, previous in zip(current_values, previous_values):
                if pd.isna(current) or pd.isna(previous) or previous == 0:
                    continue
                deviation = ((current - previous) / previous) * 100
                if _apply_operator(abs(deviation), operator, threshold):
                    failed_count += 1

        status = "P" if failed_count == 0 else "F"
        results.append({
            "Sheet Name": sheet_name,
            "Field": metric_name,
            "Test Type": "Data Validation",
            "Test Name": "Trend Validation",
            "Status (P/F)": status,
            "Failed Count": int(failed_count),
        })
    return results
