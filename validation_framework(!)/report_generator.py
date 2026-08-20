"""
report_generator.py
--------------------
Responsible ONLY for building Validation_Report.xlsx. Contains no
validation logic - it only formats and writes results that were
already computed by validations/validation.py.

Report structure (exactly three sheets):
    Sheet 1: Validation Summary   - every test executed (pass, fail,
                                     and skipped), Sheet Name / Field /
                                     Test Type / Test Name / Status /
                                     Failed Count. This includes both
                                     the dynamic Rule Determination
                                     tests and the Data Validation
                                     tests, so it doubles as a record
                                     of which rule was discovered for
                                     each column.
    Sheet 2: Failed Records       - one row per individual failing
                                     cell across the whole workbook.
    Sheet 3: Execution Summary    - run-level totals and status.
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FAIL_FILL = PatternFill(start_color="FDE9E9", end_color="FDE9E9", fill_type="solid")
PASS_FILL = PatternFill(start_color="E7F4E4", end_color="E7F4E4", fill_type="solid")
SKIP_FILL = PatternFill(start_color="FFF6E0", end_color="FFF6E0", fill_type="solid")

STATUS_FILLS = {"PASS": PASS_FILL, "FAIL": FAIL_FILL, "SKIPPED": SKIP_FILL}


def _write_header(sheet, headers):
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"


def _autofit_columns(sheet, min_width=10, max_width=60):
    for column_cells in sheet.columns:
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=0,
        )
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = max(min_width, min(length + 2, max_width))


def _add_table(sheet, name, n_rows, n_cols):
    """Turn the written range into a proper Excel Table, which gives
    the user built-in column filters for free."""
    if n_rows < 1:
        return
    last_col_letter = get_column_letter(n_cols)
    table_ref = f"A1:{last_col_letter}{n_rows + 1}"
    table = Table(displayName=name, ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    sheet.add_table(table)


# --------------------------------------------------------------------------
# Sheet 1: Validation Summary
# --------------------------------------------------------------------------

def _write_summary_sheet(workbook, summary_rows):
    sheet = workbook.active
    sheet.title = "Validation Summary"

    headers = ["Sheet Name", "Field", "Test Type", "Test Name", "Status", "Failed Count"]
    _write_header(sheet, headers)

    for row_index, row in enumerate(summary_rows, start=2):
        sheet.cell(row=row_index, column=1, value=row["Sheet Name"])
        sheet.cell(row=row_index, column=2, value=row["Field"])
        sheet.cell(row=row_index, column=3, value=row["Test Type"])
        sheet.cell(row=row_index, column=4, value=row["Test Name"])
        status_cell = sheet.cell(row=row_index, column=5, value=row["Status"])
        sheet.cell(row=row_index, column=6, value=row["Failed Count"])

        fill = STATUS_FILLS.get(row["Status"], PASS_FILL)
        for col in range(1, 7):
            sheet.cell(row=row_index, column=col).fill = fill
        status_cell.font = Font(bold=True)

    _add_table(sheet, "ValidationSummary", len(summary_rows), len(headers))
    _autofit_columns(sheet)


# --------------------------------------------------------------------------
# Sheet 2: Failed Records
# --------------------------------------------------------------------------

def _write_failed_records_sheet(workbook, detailed_failures):
    sheet = workbook.create_sheet("Failed Records")

    headers = [
        "Sheet Name", "Field", "Row", "Value",
        "Validation Type", "Required Datatype", "Error Message",
    ]
    _write_header(sheet, headers)

    if detailed_failures:
        for row_index, failure in enumerate(detailed_failures, start=2):
            sheet.cell(row=row_index, column=1, value=failure["Sheet Name"])
            sheet.cell(row=row_index, column=2, value=failure["Field"])
            sheet.cell(row=row_index, column=3, value=failure["Row"])
            sheet.cell(row=row_index, column=4, value=str(failure["Value"]))
            sheet.cell(row=row_index, column=5, value=failure["Validation Type"])
            sheet.cell(row=row_index, column=6, value=failure["Required Datatype"])
            sheet.cell(row=row_index, column=7, value=failure["Error Message"])
            for col in range(1, 8):
                sheet.cell(row=row_index, column=col).fill = FAIL_FILL

        _add_table(sheet, "FailedRecords", len(detailed_failures), len(headers))
    else:
        sheet.cell(row=2, column=1, value="No failed records found.")

    _autofit_columns(sheet)


# --------------------------------------------------------------------------
# Sheet 3: Execution Summary
# --------------------------------------------------------------------------

def _write_execution_summary_sheet(workbook, execution_stats):
    sheet = workbook.create_sheet("Execution Summary")

    headers = ["Metric", "Value"]
    _write_header(sheet, headers)

    for row_index, (label, value) in enumerate(execution_stats.items(), start=2):
        sheet.cell(row=row_index, column=1, value=label).font = Font(bold=True)
        value_cell = sheet.cell(row=row_index, column=2, value=value)

        if label == "Execution Status":
            value_cell.font = Font(bold=True)
            value_cell.fill = PASS_FILL if value == "SUCCESS" else FAIL_FILL

    _autofit_columns(sheet)


# --------------------------------------------------------------------------
# Public entry point
# --------------------------------------------------------------------------

def generate_report(config, summary_rows, detailed_failures, execution_stats, output_file):
    """
    Write the three-sheet Validation_Report.xlsx workbook.

    Parameters
    ----------
    config : dict
        The JSON configuration used for this run (currently unused by
        the report itself, kept in the signature for interface
        stability / future use).
    summary_rows : list[dict]
        Every test executed (sheet existence, rule determination,
        data validation), from validation.run_validations().
    detailed_failures : list[dict]
        One row per individual failing cell.
    execution_stats : dict
        Run-level totals for the Execution Summary sheet.
    output_file : str
        Destination path for the .xlsx report.
    """
    workbook = Workbook()

    _write_summary_sheet(workbook, summary_rows)
    _write_failed_records_sheet(workbook, detailed_failures)
    _write_execution_summary_sheet(workbook, execution_stats)

    output_dir = os.path.dirname(output_file)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    workbook.save(output_file)
